#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
年度公司事实层 (Annual Company Fact Layer) 构建
================================================
范围：IA 年度长期业务 L8-L19 公司级表（有效 inforce + 新造 new business）
  - 2024（RBC schema）：L8-L13 inforce, L14-L19 new business
  - 2023（pre-RBC schema）：同范围，列序为 年付(annual) 在前 整付(single) 在后
单位：company 表为 HKD thousand ($'000)；保单数/受保人数 为 count。
标签：年度 = certified（官方年度审计统计）。
设计：标签驱动列映射（不依赖固定列号）；保留 Market Total 控制记录；缺失标 NULL。
依据：annual_long_schema_registry_v0.1.yaml；L8-L19 公司表解析规则（12 层 test schema）。
"""
import os, glob, sqlite3, re, hashlib

ANNUAL_ROOT = "/Users/a112233/Desktop/Jasper工作文档（不含EA项目）/Jasper AI协同经验引擎/AI工程能力整改项目/FDE项目/HKIA/12_分析框架验证_Validate_Framework/01_sources/raw/SRC-REG-IA-LTA"

YEARS = {"2024": "rbc", "2023": "pre_rbc", "2022": "pre_rbc"}

# 公司表分类
INFORCE_TABLES = ["L8","L9","L10","L11","L12","L13"]
NEWBUSINESS_TABLES = ["L14","L15","L16","L17","L18","L19"]

# 表主题（用于维度和可选字段）
TABLE_SUBJECT = {
 "L8":"non_linked_individual_life","L9":"linked_individual_life","L10":"non_retirement_group",
 "L11":"retirement_scheme","L12":"annuity","L13":"total_inforce",
 "L14":"non_linked_individual_life_nb","L15":"linked_individual_life_nb","L16":"total_individual_life_nb",
 "L17":"individual_annuity_nb","L18":"non_retirement_group_nb","L19":"total_long_term_nb",
}

def sha256(path):
    h=hashlib.sha256()
    with open(path,"rb") as fh:
        for b in iter(lambda: fh.read(65536), b""): h.update(b)
    return h.hexdigest()

def cell_kind(v):
    if v is None: return None
    if isinstance(v,(int,float)): return "num"
    s=str(v).strip()
    if re.fullmatch(r"-+", s): return "num_nil"   # 占位符 "-"
    if re.fullmatch(r"N\.?A\.?", s): return "na"
    if s=="": return None
    return "str"

def norm_header_text(v):
    if v is None: return ""
    return str(v).replace("\n"," ").strip()

def detect_schema(file):
    """读 sheet 标题行判定报告年 + RBC/pre-RBC"""
    import openpyxl
    wb=openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws=wb[wb.sheetnames[0]]
    title=""
    years_zh={ "二〇二":"202", "202":"" }
    for r in range(1,6):
        for c in range(1, min(ws.max_column,12)+1):
            v=ws.cell(row=r,column=c).value
            if v: title+=" "+norm_header_text(v)
    wb.close()
    return title

def parse_company_table(file, table_id, schema, year):
    import openpyxl
    wb=openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws=wb[wb.sheetnames[0]]
    # materialize to grid for fast random access
    grid=[]
    for row in ws.iter_rows():
        grid.append([c.value for c in row])
    maxc=len(grid[0]) if grid else 0
    def G(r,c):
        if r-1 < len(grid) and c-1 < len(grid[r-1]): return grid[r-1][c-1]
        return None
    # ---- 定位表头行（含 Name of Insurer）----
    hdr_row=None
    for r in range(1, len(grid)+1):
        for c in range(2, maxc+1):
            v=G(r,c)
            if v and "Name of" in str(v):
                hdr_row=r; break
        if hdr_row: break
    if hdr_row is None:
        wb.close(); return []
    # ---- 语义列映射：hdr_row 及上方 group 行 ----
    group={}; gr=hdr_row-1
    if gr>=1:
        for c in range(2,maxc+1):
            v=G(gr,c)
            if v: group[c]=norm_header_text(v)
    def sem_for(c):
        txt="".join([group.get(c,"")," ",norm_header_text(G(hdr_row,c))])
        return txt.lower()
    col_sem={}
    for c in range(2,maxc+1):
        t=sem_for(c)
        if "name of insurer" in t or "保險公司名稱" in t: col_sem[c]="insurer_name"
        elif "financial year end" in t or "終結日期" in t or "財政年度" in t: col_sem[c]="fy_end"
        elif "no. of policies" in t or "保單數目" in t: col_sem[c]="policy_count"
        elif "no. of lives" in t or "受保人數" in t: col_sem[c]="lives"
        elif "schemes" in t and "no." in t: col_sem[c]="scheme_count"
        elif "sums assured" in t or "保額" in t: col_sem[c]="sums_assured"
        elif ("revenue premium" in t) or ("premium" in t and "premiums" in t):
            if "single" in t or "整付" in t: col_sem[c]="premium_single"
            elif "annual" in t or "年付" in t: col_sem[c]="premium_annual"
            elif "revenue premium" in t and "revenue premium" not in "":
                col_sem[c]="premium"
            else: col_sem[c]="premium"
        elif "current estimate" in t or "現時估計值" in t: col_sem[c]="current_estimate"
        elif "net liabilities" in t or "淨負債" in t: col_sem[c]="net_liabilities"
        elif "contribution" in t or "供款" in t:
            if "single" in t: col_sem[c]="contribution_single"
            else: col_sem[c]="contribution_annual"
        elif "premium" in t: col_sem[c]="premium"
        else: col_sem[c]="unknown"
    facts=[]
    data_start=hdr_row+2
    for r in range(data_start, len(grid)+1):
        name_zh=None; name_en=None; semvals={}
        for c,sem in col_sem.items():
            v=G(r,c)
            if sem=="insurer_name":
                if isinstance(v,str) and re.search(r"[一-鿿]", v): name_zh=v.strip()
                elif isinstance(v,(str,int,float)): name_en=str(v).strip()
            else:
                semvals[sem]=v
        combined=(name_zh or "")+(name_en or "")
        empty_vals=all(v in (None,) for v in semvals.values())
        if not name_zh and not name_en and empty_vals: continue
        if not name_zh and not name_en:
            # 可能备注/摘要/空尾
            if empty_vals: break
            else: continue
        if "市場總額" in combined or "market total" in combined.lower():
            for sem,v in semvals.items():
                kind=cell_kind(v)
                if kind in ("na",): continue
                if kind=="num_nil": vr=0.0
                elif kind=="num": vr=float(v)
                else: continue
                facts.append(dict(report_year=int(year), schema=schema, table_id=table_id,
                    subject=TABLE_SUBJECT[table_id], entity_scope="market_total",
                    insurer_id=None, insurer_name_source=combined.strip(), metric_sem=sem,
                    value_raw=vr, value_na=0, unit=("count" if sem in ("policy_count","lives","scheme_count") else "hkd_thousand"),
                    is_market_total=1))
            continue
        if re.search(r"^註|Note", combined): break
        ins_name=name_en or name_zh
        for sem,v in semvals.items():
            kind=cell_kind(v)
            if kind in ("na",): vr=None; fna=1
            elif kind=="num_nil": vr=0.0; fna=0
            elif kind=="num": vr=float(v); fna=0
            else: continue
            facts.append(dict(report_year=int(year), schema=schema, table_id=table_id,
                subject=TABLE_SUBJECT[table_id], entity_scope="insurer", insurer_id=None,
                insurer_name_source=ins_name, metric_sem=sem, value_raw=vr, value_na=fna,
                unit=("count" if sem in ("policy_count","lives","scheme_count") else "hkd_thousand"),
                is_market_total=0))
    wb.close()
    return facts

def main():
    out_dir=os.path.join(os.path.dirname(__file__),"..","data"); out_dir=os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    db_path=os.path.join(out_dir,"annual_company_fact_layer_2022_2024.db")
    all_facts=[]
    stat={}
    for year,schema in YEARS.items():
        table_ids = NEWBUSINESS_TABLES if True else INFORCE_TABLES  # all L8-L19
        table_ids = INFORCE_TABLES + NEWBUSINESS_TABLES
        for t in table_ids:
            fs=glob.glob(ANNUAL_ROOT+f"/{year}/full_annual_set/*{t}*")
            if not fs: 
                print(f"  [warn] no {year} {t}"); continue
            f=fs[0]
            facts=parse_company_table(f, t, schema, year)
            for fa in facts: fa["source_file"]=os.path.basename(f); fa["checksum_sha256"]=sha256(f)
            all_facts.extend(facts)
            n=len(facts); stat[(year,t)]=n
            print(f"  {year} {t}: {n} facts  ({os.path.basename(f)})")
    if os.path.exists(db_path): os.remove(db_path)
    conn=sqlite3.connect(db_path)
    cur=conn.cursor()
    cur.execute("""
      CREATE TABLE company_facts(
        report_year INTEGER, schema TEXT, table_id TEXT, subject TEXT,
        entity_scope TEXT, insurer_id TEXT, insurer_name_source TEXT,
        metric_sem TEXT, value_raw REAL, value_na INTEGER DEFAULT 0,
        unit TEXT, is_market_total INTEGER DEFAULT 0,
        source_file TEXT, checksum_sha256 TEXT)
    """)
    cur.executemany("""INSERT INTO company_facts VALUES
      (:report_year,:schema,:table_id,:subject,:entity_scope,:insurer_id,:insurer_name_source,
       :metric_sem,:value_raw,:value_na,:unit,:is_market_total,:source_file,:checksum_sha256)""",
      [dict(x, value_na=int(x.get("value_na",0))) if isinstance(x,dict) else x for x in all_facts])
    conn.commit()
    n=cur.execute("SELECT COUNT(*) FROM company_facts").fetchone()[0]
    nf=cur.execute("SELECT COUNT(*) FROM company_facts WHERE entity_scope='insurer'").fetchone()[0]
    ntot=cur.execute("SELECT COUNT(*) FROM company_facts WHERE entity_scope='market_total'").fetchone()[0]
    conn.close()
    print(f"\nDB: {db_path}")
    print(f"  total facts={n}  insurer_facts={nf}  market_total_facts={ntot}")
    print("  per(year,table):", stat)

if __name__=="__main__":
    main()
