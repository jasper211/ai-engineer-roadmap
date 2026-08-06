#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_channel_data.py —— 从 IA 原始季度文件抽取公司级×销售渠道保费数据
范围：2023Q1 / 2024Q1 / 2025Q1 / 2026Q1（个人新造长期业务）
重点：经纪渠道（Brokers）年化保费、各渠道结构与占比
输出：scripts/data/channel_data_2023_2026Q1.json

格式说明
- 旧格式(2023Q1/2024Q1) Table L1(d)：列=代理a/银行b/经纪c/直接d/其他e/总额f，
  每渠道 2 列（Single, Annualized），单位千港元。无保单数。
- 新格式(2025Q1/2026Q1) Table L1 (channel)：列分布为 保单数[代理..总额] +
  保费[代理..总额]，每渠道保费 2 列（Single, Annualized）。含保单数。
"""
import json
import os
from pathlib import Path

BASE = Path(__file__).resolve().parents[1] / "report"        # 生成页目录
SRCROOT = Path(__file__).resolve().parents[3].joinpath(
    "12_分析框架验证_Validate_Framework/01_sources/raw/SRC-REG-IA-LTQ")
OUT_DIR = Path(__file__).resolve().parent / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

QUARTERS = ["2023Q1", "2024Q1", "2025Q1", "2026Q1"]
TXLS = {  # 季度 -> 文件名
    "2023Q1": "1q23long.xls",
    "2024Q1": "1q24long.xls",
    "2025Q1": "1q25long.xlsx",
    "2026Q1": "1q26_long.xlsx",
}
OLD = {"2023Q1", "2024Q1"}   # 旧格式（.xls）

# 渠道在行列中的字段位置（保费：Single, Annualized）
# 旧格式列：代理(2,3) 银行(4,5) 经纪(6,7) 直接(8,9) 其他(10,11) 总额(12,13)
# 新格式保费列：代理(15,16) 银行(17,18) 经纪(19,20) 直接(21,22) 其他(23,24) 总额(25,26)
CH_OLD = {"agents": (2,3), "banks": (4,5), "brokers": (6,7),
          "direct": (8,9), "others": (10,11), "total": (12,13)}
CH_NEW = {"agents": (15,16), "banks": (17,18), "brokers": (19,20),
          "direct": (21,22), "others": (23,24), "total": (25,26)}

def num(v):
    """把 xlsx/xls 值规范化；'-'/'None'→None"""
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        if not v or v in ("-", "--", ""): return None
        try: return float(v.replace(",", ""))
        except: return None
    try: return float(v)
    except: return None

def old_sheet(q, ws):
    """旧格式：返回 [ {entity, channels:{渠道: (single, annualized)}} ]"""
    rows = []
    for r in range(13, ws.nrows):
        en = str(ws.cell_value(r, 0)).strip()
        if not en or en in ("-", ""): continue
        ch = {}
        for k, (c1, c2) in CH_OLD.items():
            ch[k] = (num(ws.cell_value(r, c1)), num(ws.cell_value(r, c2)))
        rows.append({"entity": en, "entity_zh": str(ws.cell_value(r,1)).strip(), "channels": ch})
    return rows

def new_sheet(q, ws):
    rows = []
    for r in range(10, ws.max_row + 1):
        en = ws.cell(r, 1).value
        if en is None: continue
        en = str(en).strip()
        if not en or en in ("-", "--", ""): continue
        # 保单数：新格式 代理(3-4)...总额(13-14)，each 整付/非整付
        pol = {}
        ch_pol = {"agents":(3,4),"banks":(5,6),"brokers":(7,8),
                  "direct":(9,10),"others":(11,12),"total":(13,14)}
        for k,(c1,c2) in ch_pol.items():
            pol[k] = (num(ws.cell(r,c1).value), num(ws.cell(r,c2).value))
        ch = {}
        for k,(c1,c2) in CH_NEW.items():
            ch[k] = (num(ws.cell(r,c1).value), num(ws.cell(r,c2).value))
        rows.append({"entity": en, "entity_zh": str(ws.cell(r,2).value or "").strip(),
                     "policies": pol, "channels": ch})
    return rows

def load_quarter(q):
    f = TXLS[q]
    path = SRCROOT / q / f
    if not path.exists():
        print(f"[跳过] {path} 不存在"); return None
    if q in OLD:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_name("Table L1(d)")
        return old_sheet(q, ws)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
        ws = wb["Table L1 (channel)"]
        return new_sheet(q, ws)

def build():
    out = {"window": "2023Q1_2026Q1_provisional", "generated_at": "2026-08",
           "metric_labels": {"single": "整付保费", "annualized": "年化保费(APE)",
                             "total_premium": "整付+年化", "policies": "保单数"},
           "quarters": {}}
    for q in QUARTERS:
        rows = load_quarter(q)
        if rows is None:
            continue
        # 汇总各渠道：总保费=单+年化；年化=APE
        agg = {}
        for ch in ["agents","banks","brokers","direct","others","total"]:
            agg[ch] = {"sum_single":0.0,"sum_annualized":0.0,"sum_total":0.0}
        for r in rows:
            for ch,(s,a) in r["channels"].items():
                if s is None and a is None: continue
                sg = s or 0.0; an = a or 0.0
                agg[ch]["sum_single"] += sg
                agg[ch]["sum_annualized"] += an
                agg[ch]["sum_total"] += sg + an
        out["quarters"][q] = {
            "n_insurers": len(rows),
            "source_file": TXLS[q],
            "channel_agg": agg,
            "insurers": rows,
        }
    return out

def save(d):
    p = OUT_DIR / "channel_data_2023_2026Q1.json"
    p.write_text(json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"已写入 {p}")
    # 摘要
    for q, qd in d["quarters"].items():
        a = qd["channel_agg"]
        tot = a["total"]["sum_total"]
        def pct(k): return (a[k]["sum_total"]/tot*100) if tot else 0
        print(f"\n{q}（{qd['n_insurers']}家） 渠道总保费占比："
              f"代理{pct('agents'):.1f}% 银行{pct('banks'):.1f}% "
              f"经纪{pct('brokers'):.1f}% 直接{pct('direct'):.1f}% 其他{pct('others'):.1f}%")
        print(f"  经纪渠道: 整付{a['brokers']['sum_single']/1e6:.2f}亿 + "
              f"年化{a['brokers']['sum_annualized']/1e6:.2f}亿 = "
              f"总{a['brokers']['sum_total']/1e6:.2f}亿")

if __name__ == "__main__":
    save(build())
