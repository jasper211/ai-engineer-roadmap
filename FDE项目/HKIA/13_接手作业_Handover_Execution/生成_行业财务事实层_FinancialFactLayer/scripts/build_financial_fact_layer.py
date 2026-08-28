#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B4 · 行业财务事实层 (Industry Financial Fact Layer) 构建脚本
=============================================================
范围：IA 季度临时财务资料 (Industry Financial Info)
  - "By Fund" sheet：行业总计/长期/分红长期/一般 四 scope，17 个资产/负债科目
  - 覆盖期：2024Q4, 2025Q1, 2025Q3, 2025Q4, 2026Q1（2025Q2 源为 OLE2 旧格式，本环境无转换工具，登记为缺口）
  - 单位：港币百万元 (HK$'million) —— 官方源直接值，不换算
  - 标签：provisional / unaudited（官方 note(4)）
幂等可重建：每次运行时清空重写 financial_fact_layer.db 的 facts 表。

依据：
  - asset: AST-IA-FINANCIAL-2026Q1-XLSX (asset_registry_phase_b_v0.1.yaml)
  - batch: ia_acquisition_batch_p0_v0.1.yaml (ASTSET-IA-FINANCIAL-2024Q4-2026Q1-XLSX)
  - coverage: ia_asset_metric_coverage_matrix_v0.2.yaml (theme T01/T31/T32)
"""
import os, sys, sqlite3, hashlib, json, datetime

RAW_ROOT = "/Users/a112233/Desktop/Jasper工作文档（不含EA项目）/Jasper AI协同经验引擎/AI工程能力整改项目/FDE项目/HKIA/12_分析框架验证_Validate_Framework/01_sources/raw/SRC-REG-IA-FINANCIAL"

# 期 -> 源文件（5 期可读 xlsx；2025Q2 为 OLE2，登记缺口）
PERIOD_FILES = {
    "2024Q4": f"{RAW_ROOT}/2024/4q24_Industry_Financial_Info.xlsx",
    "2025Q1": f"{RAW_ROOT}/2025/1q25_Industry_Financial_Info.xlsx",
    "2025Q3": f"{RAW_ROOT}/2025/3q25_Industry_Financial_Info.xlsx",
    "2025Q4": f"{RAW_ROOT}/2025/4q25_IndustryFinancial_Info.xlsx",
    "2026Q1": f"{RAW_ROOT}/2026/1q26_Industry_Financial_Info.xlsx",
}
PERIOD_GAP = {
    "2025Q2": {
        "reason": "源文件为 OLE2 Composite Document (旧 .xls 伪装 .xlsx)，本环境无 LibreOffice/olefile/msoffcrypto，xlrd 无法定位 Workbook stream 解析失败",
        "normalized_copy": "12_分析框架验证_Validate_Framework/04_normalized/legacy_excel_converted/2q25_Industry_Financial_Info.xlsx",
        "normalized_status": "乱码不可用（OLE2 转换产物损坏）",
        "status": "pending_conversion_tool"
    }
}

# "By Fund" sheet：row4=fund scopes(col2-5), rows5-21=17 科目
FUND_SHEET = "By Fund"
SCOPE_COLS = {  # column -> scope_key
    2: "industry_total",
    3: "long_term",
    4: "participating_long_term",
    5: "general_business",
}
# rows5-21 科目顺序（标签已验证与全部可读文件一致）
ITEM_ROWS = list(range(5, 22))  # 5..21
ITEM_LABEL = {
    5: "total_assets",
    6: "cash_and_deposits",
    7: "debt_securities",
    8: "equities_portfolio",
    9: "properties",
    10: "loans_and_advances",
    11: "unit_linked_or_retirement_policyholder_assets",
    12: "other_financial_assets",
    13: "reinsurance_assets",
    14: "tax_assets",
    15: "other_assets",
    16: "total_liabilities",
    17: "insurance_liabilities_incl_reinsurance",
    18: "financial_liabilities",
    19: "tax_liabilities",
    20: "other_liabilities",
    21: "net_assets",
}

COUNT_UNIT = "HKD_million"
FLAG = "provisional_unaudited"


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for blk in iter(lambda: fh.read(65536), b""):
            h.update(blk)
    return h.hexdigest()


def parse_period_file(period, path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[FUND_SHEET]
    facts = []
    for r in ITEM_ROWS:
        item_id = ITEM_LABEL[r]
        raw_label = ws.cell(row=r, column=1).value
        label_zh = (str(raw_label).split("\n")[0] if raw_label else "")
        for col, scope in SCOPE_COLS.items():
            v = ws.cell(row=r, column=col).value
            v = None if v is None else float(v)
            facts.append({
                "period": period,
                "fund_scope": scope,
                "item_id": item_id,
                "item_label_zh": label_zh,
                "value_hkd_million": v,
                "unit": COUNT_UNIT,
                "flag": FLAG,
                "source_file": os.path.basename(path),
                "checksum_sha256": sha256(path),
            })
    wb.close()
    return facts


def main():
    out_dir = os.path.join(os.path.dirname(__file__), "..", "data")
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    db_path = os.path.join(out_dir, "financial_fact_layer.db")

    all_facts = []
    parsed_periods = []
    for period, path in PERIOD_FILES.items():
        if not os.path.exists(path):
            print(f"[warn] missing source for {period}: {path}")
            continue
        facts = parse_period_file(period, path)
        all_facts.extend(facts)
        parsed_periods.append(period)
        print(f"  parsed {period}: {len(facts)} facts")

    print("parsed_periods:", parsed_periods)
    print("total facts:", len(all_facts))

    # write DB
    if os.path.exists(db_path):
        os.remove(db_path)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE financial_facts (
            period TEXT,
            fund_scope TEXT,
            item_id TEXT,
            item_label_zh TEXT,
            value_hkd_million REAL,
            unit TEXT,
            flag TEXT,
            source_file TEXT,
            checksum_sha256 TEXT
        )
    """)
    cur.executemany("""
        INSERT INTO financial_facts VALUES
        (:period,:fund_scope,:item_id,:item_label_zh,:value_hkd_million,:unit,:flag,:source_file,:checksum_sha256)
    """, all_facts)
    conn.commit()

    # summary counts
    n_period = cur.execute("SELECT COUNT(DISTINCT period) FROM financial_facts").fetchone()[0]
    n_scope = cur.execute("SELECT COUNT(DISTINCT fund_scope) FROM financial_facts").fetchone()[0]
    n_item = cur.execute("SELECT COUNT(DISTINCT item_id) FROM financial_facts").fetchone()[0]
    conn.close()

    print(f"\nDB written: {db_path}")
    print(f"  distinct periods: {n_period}")
    print(f"  distinct fund_scopes: {n_scope}")
    print(f"  distinct item_ids: {n_item}")
    print(f"  distinct labels (zh): {len(set(f['item_label_zh'] for f in all_facts))}")
    print("\nGAP_PENDING (2025Q2):", PERIOD_GAP)


if __name__ == "__main__":
    main()
