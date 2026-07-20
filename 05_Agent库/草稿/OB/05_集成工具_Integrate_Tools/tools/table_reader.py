#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工具：把 03_发布成果-交付物 下的 xlsx/csv 权威数据表，读成可以喂给LLM做
知识原子提炼的结构化文本块。跟 concept_note_extraction.py 的自由文本文档
不是一回事——这里的输入本来就是结构化表格（价值节点清单/KPI映射表等），
不需要判断"这段话算不算一个原子"，需要判断的是"哪些行该合并成一个原子、
该怎么用自然语言转述"。

真实探查过 D1_价值节点清单/流程库状态监控表 这两份文件，发现：
- 都是多sheet工作簿（10-14个sheet），不是单表
- 大部分sheet是真数据（节点详情卡/熔断清单/未决裁定项/变更日志等），
  只有"0.方法论与说明"/"填写说明"这类是纯元信息，不该逐行提炼
- header行位置不固定（0-3行banner不等），得动态探测，不能假设第0行就是表头
- 同一张逻辑表在文件夹里堆了多个历史版本（V3.38~V3.44共7份），只该取
  最新版本，否则会把同样内容重复提炼7遍
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

VERSION_RE = re.compile(r"^(.*?)_[Vv](\d+(?:\.\d+)*)(\.\w+)$")

# sheet名命中这些关键字的，判定为"元信息/说明"性质，不逐行提炼数据
META_SHEET_KEYWORDS = ("方法论", "说明", "填写", "readme", "index", "目录")

MAX_HEADER_SCAN_ROWS = 6


def group_latest_versions(files: List[Path]) -> List[Path]:
    """同一张逻辑表的多个历史版本文件，只保留版本号最大的那份。文件名对不上
    版本命名模式（如 dim_kpi_v3.3_权威层.csv，版本号后面还带后缀）的，按"没法
    分组、独立成组"处理，直接保留——宁可漏判成多个独立文件，也不因为分组
    出错误删了本该保留的文件。"""
    groups: Dict[str, List[Tuple[tuple, Path]]] = {}
    standalone: List[Path] = []
    for f in files:
        m = VERSION_RE.match(f.name)
        if not m:
            standalone.append(f)
            continue
        base, ver, ext = m.groups()
        ver_tuple = tuple(int(x) for x in ver.split("."))
        groups.setdefault(base + ext, []).append((ver_tuple, f))

    latest = [sorted(items)[-1][1] for items in groups.values()]
    return latest + standalone


def is_meta_sheet(sheet_name: str) -> bool:
    lname = sheet_name.lower()
    return any(kw.lower() in lname for kw in META_SHEET_KEYWORDS)


def _detect_header_row(rows: List[tuple]) -> int:
    """在前MAX_HEADER_SCAN_ROWS行里，找非空格子占比最高的一行当表头——
    真实数据显示banner/标题行通常只有1-2个非空格子（合并单元格的标题文本），
    真正的表头行几乎每一列都有值（列名）。命中率打平时取更靠前的行。"""
    best_idx, best_fill = 0, -1
    for i, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        fill = sum(1 for c in row if c is not None and str(c).strip() != "")
        if fill > best_fill:
            best_fill, best_idx = fill, i
    return best_idx


def read_table_blocks(file_path: Path) -> List[Dict]:
    """读一份xlsx/csv，返回该文件里"值得逐行提炼"的数据块列表——每个xlsx的
    每个非元信息sheet算一块，csv整份算一块。每块含 {source_label, columns, rows}。
    rows是list[dict]（列名->值），已经跳过表头之上的banner行和全空行。"""
    blocks = []
    if file_path.suffix.lower() == ".csv":
        try:
            df = pd.read_csv(file_path, encoding="utf-8-sig")
        except pd.errors.ParserError:
            # 真实复现过：D3_L3调整记录里"调整原因"这类自由文本字段包含未加
            # 引号的逗号，导致该行字段数对不上表头，pandas C引擎直接整份报错。
            # 一行格式错误不该让整份224行的表都读不出来——退回python引擎+
            # on_bad_lines跳过坏行，宁可丢1行也不丢整份文件。
            try:
                df = pd.read_csv(file_path, encoding="utf-8-sig", engine="python", on_bad_lines="skip")
            except Exception:
                return blocks
        except Exception:
            return blocks
        df = df.dropna(how="all")
        if df.empty:
            return blocks
        blocks.append({
            "source_label": file_path.name,
            "columns": [str(c) for c in df.columns],
            "rows": df.to_dict(orient="records"),
        })
        return blocks

    if file_path.suffix.lower() != ".xlsx":
        return blocks

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if is_meta_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue
        header_idx = _detect_header_row(all_rows)
        header = [str(c).strip() if c else f"列{i}" for i, c in enumerate(all_rows[header_idx])]
        data_rows = []
        for row in all_rows[header_idx + 1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            row_dict = {header[i]: row[i] for i in range(min(len(header), len(row)))}
            data_rows.append(row_dict)
        if not data_rows:
            continue
        blocks.append({
            "source_label": f"{file_path.name} / {sheet_name}",
            "columns": header,
            "rows": data_rows,
        })
    wb.close()
    return blocks


def serialize_rows_markdown(columns: List[str], rows: List[Dict]) -> str:
    """把一批行序列化成markdown表格文本，喂给LLM。None/NaN值转成空字符串，
    不留Python的'None'/'nan'字面量污染LLM看到的内容。"""
    def clean(v):
        if v is None:
            return ""
        s = str(v)
        return "" if s.lower() == "nan" else s

    lines = ["| " + " | ".join(columns) + " |", "| " + " | ".join(["---"] * len(columns)) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(clean(row.get(c)) for c in columns) + " |")
    return "\n".join(lines)


def list_table_candidates(project_root: str, subdir: str = "03_发布成果-交付物") -> List[Path]:
    """在project_root/subdir下找所有xlsx/csv（跳过_归档目录），按版本分组
    只保留最新版本。"""
    base = Path(project_root) / subdir
    if not base.exists():
        return []
    all_files = [
        p for p in base.rglob("*")
        if p.suffix.lower() in (".xlsx", ".csv") and "_归档" not in p.parts
        and not p.name.startswith("~$")  # Excel临时锁文件
    ]
    return group_latest_versions(all_files)
