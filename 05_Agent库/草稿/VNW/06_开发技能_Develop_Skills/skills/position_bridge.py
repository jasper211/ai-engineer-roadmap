"""L4 → 候选Agent(旧24/30个聚合口径) → 岗位族 → 实际在岗人力 桥接。

三份源材料，2026-08-04确认可以拼接：
1. `候选Agent目录_数据表_v3.xlsx`的"L4明细"sheet(368条，L4编码→候选Agent名称)
2. `候选Agent岗位族映射_HR对齐版_v1.0.csv`(32条，候选Agent→岗位族，本次会话按
   HR权威文档`2026-07-20_68L3岗位族归属设计_v6.1_SUBMITTED.md`对齐过冲突)
3. `2026-07-28_人员归属方案_v5.3_SUBMITTED.md`一节"映射统计总表"的岗位族级
   实际在岗人数(逐人可查，文档内部有跨文件平衡表校验，差异=0)

这条链路解决的是`dim_agent.owner_position_family`361条全为空、走不通的问题——
不是从`dim_agent`出发，是从旧candidate agent口径的L4明细出发，因为L4编码是
两套体系共享的稳定key。
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

UNASSIGNED_MARKER = "⚠️未覆盖(待归口)"

# 岗位族实际在岗人数，来源：2026-07-28_人员归属方案_v5.3_SUBMITTED.md
# 第一节"映射统计总表"，文档内部逐人差异表(第九节)校验差异=0，可信。
POSITION_FAMILY_HEADCOUNT = {
    "JF-01": {"name": "保司战略族", "headcount": 2},
    "JF-02": {"name": "保司关系族", "headcount": 0},
    "JF-03": {"name": "机构业务族", "headcount": 52},
    "JF-04": {"name": "事业部运营族", "headcount": 39},
    "JF-05": {"name": "理财师发展族", "headcount": 1},
    "JF-06": {"name": "权益服务族", "headcount": 11},
    "JF-07": {"name": "佣金合规族", "headcount": 5},
    "职能层": {"name": "职能支撑层", "headcount": 26},
}
HEADCOUNT_SOURCE = "2026-07-28_人员归属方案_v5.3_SUBMITTED.md 一节映射统计总表"


def load_l4_to_candidate_agent(xlsx_path: Path) -> dict[str, str]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["L4明细"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    l4_idx = headers.index("L4编码")
    agent_idx = headers.index("候选Agent(30个)")
    result = {}
    for row in rows:
        l4_code = str(row[l4_idx] or "").strip().upper()
        agent = str(row[agent_idx] or "").strip()
        if l4_code and agent and agent != UNASSIGNED_MARKER:
            result[l4_code] = agent
    return result


def load_candidate_agent_to_position_family(csv_path: Path) -> dict[str, dict]:
    result = {}
    with Path(csv_path).open(encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            agent = row.get("候选Agent", "").strip()
            family_code = row.get("岗位族代码", "").strip()
            if not agent or not family_code:
                continue
            result[agent] = {
                "family_code": family_code,
                "family_name": row.get("岗位族名称", "").strip(),
                "confidence": row.get("数据可信度", "").strip(),
            }
    return result


def build_l4_position_bridge(
    candidate_agent_xlsx: Path, hr_aligned_csv: Path
) -> dict[str, dict]:
    """返回 {l4_code: {candidate_agent, family_code, family_name, headcount,
    confidence, headcount_source}}。只收录能完整走完三段桥接的L4；缺一段的
    如实不收录，不用0或空字符串伪装成"查过但没有"。"""
    l4_to_agent = load_l4_to_candidate_agent(candidate_agent_xlsx)
    agent_to_family = load_candidate_agent_to_position_family(hr_aligned_csv)

    result = {}
    for l4_code, agent in l4_to_agent.items():
        family_info = agent_to_family.get(agent)
        if family_info is None or family_info["family_code"] == "待确认":
            continue
        headcount_info = POSITION_FAMILY_HEADCOUNT.get(family_info["family_code"])
        if headcount_info is None:
            continue
        result[l4_code] = {
            "candidate_agent": agent,
            "family_code": family_info["family_code"],
            "family_name": family_info["family_name"],
            "confidence": family_info["confidence"],
            "headcount": headcount_info["headcount"],
            "headcount_source": HEADCOUNT_SOURCE,
        }
    return result
