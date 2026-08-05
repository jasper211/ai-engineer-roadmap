"""基于权威库和合格补充证据构建L3基础模型快照。"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from tools.evidence import EvidenceClass, EvidenceRecord, EvidenceStatus, SourceRef, authoritative
from tools.postgres_reader import PostgresL3Reader
from tools.snapshot_writer import write_snapshot
from skills.blueprint_parser import parse_blueprint
from skills.l3_analysis_contract import (
    analysis_input_hash,
    build_analysis_envelope,
    eligible_analysis_evidence_ids,
    validate_analysis_package,
)
from skills.unified_analysis import l3_synthesis

VALID_TIERS = {"Human", "Aug", "Hybrid", "Auto"}
D_FIELDS = (
    "agent_d1_input_struct",
    "agent_d2_rule_clear",
    "agent_d3_output_verify",
    "agent_d4_api_reach",
    "agent_d5_fallback",
    "agent_d6_compliance",
)


@dataclass(frozen=True)
class BlueprintIndex:
    l3_code: str
    l3_name: str
    version: str
    filename: str


def load_blueprint_index(path: Path) -> dict[str, BlueprintIndex]:
    result = {}
    with Path(path).open(encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            code = row.get("l3_code", "").strip()
            if not code:
                continue
            full_code = (code if code.startswith("L3-") else f"L3-{code}").upper()
            raw_version = row.get("version", "").strip()
            version_match = re.search(r"(\d+)\D+(\d+)", raw_version)
            normalized_version = f"V{version_match.group(1)}.{version_match.group(2)}" if version_match else raw_version
            result[full_code] = BlueprintIndex(
                l3_code=full_code,
                l3_name=row.get("l3_name", "").strip(),
                version=normalized_version,
                filename=row.get("filename", "").strip(),
            )
    return result


BLUEPRINT_FILENAME_RE = re.compile(r"^流程蓝图_([A-Za-z0-9\-]+)_(.+)_V([\d.]+)\.md$")


def load_blueprint_index_from_dir(dir_path: Path) -> dict[str, BlueprintIndex]:
    """按实际磁盘文件判定蓝图覆盖，取代手工维护的`L3蓝图覆盖清单_v1.0.csv`。

    2026-07-29核实：该CSV索引67个L3里有18个缺失记录，但其中16个蓝图文件
    实际存在——索引本身过期，不是内容缺口。直接扫描`blueprint_dir`（EA项目
    真实L3流程库目录）按文件名解析编码/名称/版本，同编码多版本时只取最大
    版本号，天然不会过期。
    """
    groups: dict[str, tuple[tuple, BlueprintIndex]] = {}
    for f in Path(dir_path).glob("流程蓝图_*.md"):
        m = BLUEPRINT_FILENAME_RE.match(f.name)
        if not m:
            continue
        code = m.group(1).upper()
        name = m.group(2)
        version = m.group(3)
        ver_tuple = tuple(int(x) for x in version.split("."))
        entry = BlueprintIndex(l3_code=code, l3_name=name, version=f"V{version}", filename=f.name)
        existing = groups.get(code)
        if existing is None or ver_tuple > existing[0]:
            groups[code] = (ver_tuple, entry)
    return {code: entry for code, (_, entry) in groups.items()}


D1D6_CSV_FIELD_MAP = {
    "agent_d1_input_struct": "D1_输入结构化",
    "agent_d2_rule_clear": "D2_规则清晰",
    "agent_d3_output_verify": "D3_输出可验",
    "agent_d4_api_reach": "D4_API可达",
    "agent_d5_fallback": "D5_降级可用",
    "agent_d6_compliance": "D6_合规可编码",
}


def d1d6_name_key(l3_code: str, l4_name: str) -> str:
    """历史L4编码迁移时使用的保守桥接键：同一L3 + 完全相同的L4名称。"""
    return f"NAME::{l3_code.strip().upper()}::{l4_name.strip()}"


def load_d1d6_supplement(csv_path: Path) -> dict[str, dict]:
    """加载`L4两阶段复核_全量368条_合并版_v1.0.csv`，按l4_code建索引。

    2026-07-29核实：`dim_process`里366条L4只有35条有D1-D6分数，此前误判为
    "只做了6个L3的评分"；实际这份CSV对368条L4全部打了分。数据库缺的是
    同步，不是打分工作本身。仅当一行6个维度全部有值时才收录，避免半截
    数据混进补充证据。
    """
    result: dict[str, dict] = {}
    name_candidates: dict[str, list[dict]] = {}
    with Path(csv_path).open(encoding="utf-8-sig") as file:
        for source_row, row in enumerate(csv.DictReader(file), start=2):
            l4_code = row.get("L4编码", "").strip()
            if not l4_code:
                continue
            values = {}
            complete = True
            for db_field, csv_col in D1D6_CSV_FIELD_MAP.items():
                raw = row.get(csv_col, "").strip()
                if raw.isdigit():
                    values[db_field] = int(raw)
                else:
                    complete = False
                    break
            if complete:
                values["_source_l4_code"] = l4_code
                values["_source_row"] = source_row
                result[l4_code] = values
                l3_code = row.get("L3编码", "").strip().upper()
                l4_name = row.get("L4名称", "").strip()
                if l3_code and l4_name:
                    name_candidates.setdefault(d1d6_name_key(l3_code, l4_name), []).append(values)
    # 只为名称在同一L3内唯一的记录建立别名索引，避免同名交付物误连。
    for key, candidates in name_candidates.items():
        if len(candidates) == 1:
            result[key] = candidates[0]
    return result


def load_skill_feasibility(xlsx_path: Path) -> dict[str, dict]:
    """加载L4 Skill封装可行性复核，保持其与自动化Tier为两条独立分析轴。

    直接编码和“同一L3+完全相同L4名称”各建一个索引；后者只在名称唯一时生效，
    用于兼容UCA等历史L4编码。SFC/SPE六条口头确认例外保留为待佐证共识层。
    """
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["L4明细_最终确认版"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    result: dict[str, dict] = {}
    name_candidates: dict[str, list[dict]] = {}
    for row_number, values in enumerate(rows, 2):
        row = dict(zip(headers, values))
        l4_code = str(row.get("L4编码") or "").strip().upper()
        l3_code = str(row.get("L3编码") or "").strip().upper()
        l4_name = str(row.get("L4活动") or "").strip()
        if not l4_code or not l3_code or not l4_name:
            continue
        if l4_code in result:
            raise ValueError(f"Skill封装评估L4编码重复：{l4_code}")
        provisional = l3_code in {"L3-SFC", "L3-SPE"}
        item = {
            "action_nature": str(row.get("动作性质") or "").strip(),
            "action_singularity": str(row.get("动作单一性") or "").strip(),
            "grade": str(row.get("最终档位") or "").strip(),
            "judgment_basis": str(row.get("判断依据") or "").strip(),
            "funds_safety_hard_gate": str(row.get("资金安全强制关卡(新增)") or "").strip() == "是",
            "physical_execution": str(row.get("物理执行类(新增)") or "").strip() == "是",
            "verification_status": "PROVISIONAL" if provisional else "VERIFIED",
            "_source_l4_code": l4_code,
            "_source_row": row_number,
        }
        result[l4_code] = item
        name_candidates.setdefault(d1d6_name_key(l3_code, l4_name), []).append(item)
    for key, candidates in name_candidates.items():
        if len(candidates) == 1:
            result[key] = candidates[0]
    return result


def skill_recommendation(assessment: dict, tier: str) -> str:
    """根据已批准的双轴规则生成可审计的设计路径，不生成优先级坐标。"""
    if assessment.get("verification_status") != "VERIFIED":
        return "待补书面佐证，暂不进入正式Skill优先级判断"
    grade = str(assessment.get("grade", ""))
    if grade.startswith("A-") and tier in {"Auto", "Aug"}:
        return "优先进入Skill原型候选池"
    if grade.startswith("A-") and tier == "Hybrid":
        return "设计为Skill执行、人工判断或批准"
    if grade.startswith("B-"):
        return "先补报告模板、规则固化或系统接口，再验证Skill"
    if grade.startswith("C-"):
        return "优先考虑Agent辅助或决策支持，不直接封装确定性Skill"
    if grade.startswith("F-"):
        return "不封装物理动作，只优化动作前后的信息流"
    return "待负责人结合任务拆分确认"


def load_analysis_packages(dir_path: Path) -> dict[str, dict]:
    """加载已经过校验/复核的L3分析包。

    同一L3可同时保留历史人工复核包和新发布模型包。新模型包已经通过
    AnalysisRunner门禁，因此优先于``reviewed``回退包；同优先级重复才
    视为配置冲突。
    """
    result = {}
    selected_priority: dict[str, int] = {}
    selected_path: dict[str, Path] = {}
    for path in Path(dir_path).glob("L3-*.json"):
        package = json.loads(path.read_text(encoding="utf-8"))
        code = path.name.split(".", 1)[0]
        priority = 2 if path.name.endswith(".model.json") else 1
        if code in result and priority == selected_priority[code]:
            raise ValueError(
                f"同一L3存在多个同优先级分析包："
                f"{selected_path[code].name}, {path.name}"
            )
        if code in result and priority < selected_priority[code]:
            continue
        result[code] = package
        selected_priority[code] = priority
        selected_path[code] = path
    return result


def load_sop_records(csv_path: Path, sop_dir: Path) -> list[dict]:
    """读取SOP生产清单，并只保留能够定位到真实文件的记录。"""
    result = []
    with Path(csv_path).open(encoding="utf-8-sig") as file:
        for source_row, row in enumerate(csv.DictReader(file), start=2):
            filename = row.get("sop_file", "").strip()
            source_path = Path(sop_dir) / filename
            if not filename or not source_path.exists():
                continue
            content = source_path.read_text(encoding="utf-8")
            result.append({
                **row,
                "_source_row": source_row,
                "source_path": str(source_path),
                "l4_codes": sorted(set(re.findall(r"\bL4-[A-Z0-9]+(?:-[A-Z0-9]+)+\b", content, flags=re.I))),
            })
    return result


def load_rule_records(csv_path: Path) -> list[dict]:
    """读取当前有效规则清单；空规则动作不作为建模依据。"""
    result = []
    with Path(csv_path).open(encoding="utf-8-sig") as file:
        for source_row, row in enumerate(csv.DictReader(file), start=2):
            if (
                row.get("rule_id", "").strip()
                and row.get("node_id", "").strip()
                and row.get("rule_action", "").strip()
            ):
                result.append({**row, "_source_row": source_row})
    return result


def load_prepared_analysis_codes(dir_path: Path) -> set[str]:
    """识别已准备或模型已返回但尚未发布的运行包，防止重复推荐。"""
    result = set()
    for request_path in Path(dir_path).glob("*/request.json"):
        request = json.loads(request_path.read_text(encoding="utf-8"))
        if request.get("status") in {"PREPARED", "MODEL_RETURNED"}:
            result.add(request.get("l3_code", ""))
    return set(filter(None, result))


def expand_l4_mapping(raw_code: str, valid_codes: set[str]) -> set[str]:
    """展开`L4-COM-01/02`一类桥接值，只返回当前L3的真实L4编码。"""
    raw = (raw_code or "").strip().upper()
    if raw in valid_codes:
        return {raw}
    parts = raw.split("/")
    if len(parts) <= 1:
        return set()
    first = parts[0]
    prefix = first.rsplit("-", 1)[0] + "-"
    candidates = {first, *(part if part.startswith("L4-") else prefix + part for part in parts[1:])}
    return candidates & valid_codes


def gate_result(status: str, checks: list[dict]) -> dict:
    return {"status": status, "checks": checks}


class L3ModelBuilder:
    schema_version = "vnw.l3-model.v1"

    def __init__(
        self,
        reader: PostgresL3Reader,
        blueprint_index: dict[str, BlueprintIndex],
        blueprint_dir: Path | None = None,
        d1d6_supplement: dict[str, dict] | None = None,
        skill_feasibility: dict[str, dict] | None = None,
        demo_registry: dict[str, str] | None = None,
        analysis_packages: dict[str, dict] | None = None,
        sop_records: list[dict] | None = None,
        rule_records: list[dict] | None = None,
        prepared_analysis_codes: set[str] | None = None,
        l3_position_category: dict[str, dict] | None = None,
        business_table_map: dict[str, list[dict]] | None = None,
        business_table_counts: dict[str, int] | None = None,
        analysis_confirmations_dir: Path | None = None,
    ):
        self.reader = reader
        self.blueprint_index = blueprint_index
        self.blueprint_dir = blueprint_dir
        self.d1d6_supplement = d1d6_supplement or {}
        self.skill_feasibility = skill_feasibility or {}
        self.demo_registry = demo_registry or {}
        self.analysis_packages = analysis_packages or {}
        self.sop_records = sop_records or []
        self.rule_records = rule_records or []
        self.prepared_analysis_codes = prepared_analysis_codes or set()
        self.l3_position_category = l3_position_category or {}
        self.business_table_map = business_table_map or {}
        self.business_table_counts = business_table_counts or {}
        self.analysis_confirmations_dir = analysis_confirmations_dir

    def build(self, l3_code: str, supplemental: list[EvidenceRecord] | None = None) -> dict:
        processes = self.reader.processes(l3_code)
        nodes = self.reader.value_nodes(l3_code)
        mappings = self.reader.vn_l4_mappings(l3_code)
        l2s = self.reader.l2_mappings(l3_code)
        kpis = self.reader.kpi_mappings(l3_code)
        value_stages = self.reader.value_stream_mappings(l3_code)
        blueprint = self.blueprint_index.get(l3_code)
        parsed_blueprint = None
        if blueprint and self.blueprint_dir:
            source_path = self.blueprint_dir / blueprint.filename
            if source_path.exists():
                parsed_blueprint = parse_blueprint(
                    source_path,
                    {row.get("l4_code", "") for row in processes if row.get("l4_code")},
                )
        evidence: dict[str, dict] = {}

        def add(record: EvidenceRecord) -> str:
            evidence[record.evidence_id] = record.to_dict()
            return record.evidence_id

        if parsed_blueprint and parsed_blueprint["structure_status"] == "PARSED":
            for step in parsed_blueprint["steps"]:
                step["evidence_ref"] = add(EvidenceRecord(
                    field_name=f"blueprint.steps.{step['step_id']}",
                    value={"step_name": step["step_name"], "l4_codes": step["l4_codes"]},
                    evidence_class=EvidenceClass.SUPPLEMENTAL,
                    status=EvidenceStatus.ACTIVE,
                    source=SourceRef(
                        source_system="L3流程库",
                        source_object=blueprint.filename,
                        source_key=f"line:{step['source_line']}",
                        source_field="step",
                        source_version=blueprint.version,
                    ),
                ))
            for decision in parsed_blueprint["decisions"]:
                decision["evidence_ref"] = add(EvidenceRecord(
                    field_name=f"blueprint.decisions.{decision['decision_id']}",
                    value={"question": decision["question"], "branches": decision["branches"]},
                    evidence_class=EvidenceClass.SUPPLEMENTAL,
                    status=EvidenceStatus.ACTIVE,
                    source=SourceRef(
                        source_system="L3流程库",
                        source_object=blueprint.filename,
                        source_key=f"line:{decision['source_line']}",
                        source_field="decision",
                        source_version=blueprint.version,
                    ),
                ))
            for node in parsed_blueprint["blueprint_value_nodes"]:
                node["evidence_ref"] = add(EvidenceRecord(
                    field_name=f"blueprint.value_nodes.{node['vn_id']}",
                    value={"vn_name": node["vn_name"], "l4_codes": node["l4_codes"], "status": node["status_text"]},
                    evidence_class=EvidenceClass.SUPPLEMENTAL,
                    status=EvidenceStatus.ACTIVE,
                    source=SourceRef(
                        source_system="L3流程库",
                        source_object=blueprint.filename,
                        source_key=f"line:{node['source_line']}",
                        source_field="value_node_mapping",
                        source_version=blueprint.version,
                    ),
                ))
            for row in parsed_blueprint["raci"]:
                row["evidence_ref"] = add(EvidenceRecord(
                    field_name=f"blueprint.raci.{row['l4_code']}",
                    value={key: row[key] for key in ("accountable", "responsible", "consulted", "informed")},
                    evidence_class=EvidenceClass.SUPPLEMENTAL,
                    status=EvidenceStatus.ACTIVE,
                    source=SourceRef(
                        source_system="L3流程库",
                        source_object=blueprint.filename,
                        source_key=f"line:{row['source_line']}",
                        source_field="raci",
                        source_version=blueprint.version,
                    ),
                ))

        l4s = []
        for row in processes:
            code = row.get("l4_code", "")
            refs = {
                field: add(authoritative(field, row.get(field), "dim_process", code, field))
                for field in (
                    "l4_name", "l4_deliverable", "l4_deliverable_type",
                    "agentifiability", "agent_human_touchpoint",
                )
            }
            d1_d6 = {}
            supplement = self.d1d6_supplement.get(code)
            if supplement is None:
                supplement = self.d1d6_supplement.get(
                    d1d6_name_key(l3_code, row.get("l4_name", ""))
                )
            for field in D_FIELDS:
                db_value = row.get(field)
                if db_value is not None:
                    d1_d6[field] = db_value
                    refs[field] = add(authoritative(field, db_value, "dim_process", code, field))
                elif supplement is not None:
                    d1_d6[field] = supplement[field]
                    refs[field] = add(EvidenceRecord(
                        field_name=field,
                        value=supplement[field],
                        evidence_class=EvidenceClass.SUPPLEMENTAL,
                        status=EvidenceStatus.ACTIVE,
                        source=SourceRef(
                            source_system="EA项目_权威数据",
                            source_object="L4两阶段复核_全量368条_合并版_v1.0.csv",
                            source_key=supplement.get("_source_l4_code", code),
                            source_field=field,
                            source_version="v1.0",
                        ),
                    ))
                else:
                    d1_d6[field] = None
                    refs[field] = add(authoritative(field, None, "dim_process", code, field))
            skill = self.skill_feasibility.get(code)
            if skill is None:
                skill = self.skill_feasibility.get(d1d6_name_key(l3_code, row.get("l4_name", "")))
            skill_payload = None
            if skill is not None:
                source_code = skill.get("_source_l4_code", code)
                source_row = skill.get("_source_row", "")
                is_provisional = skill.get("verification_status") == "PROVISIONAL"
                skill_payload = {
                    key: skill[key] for key in (
                        "action_nature", "action_singularity", "grade", "judgment_basis",
                        "funds_safety_hard_gate", "physical_execution", "verification_status",
                    )
                }
                skill_payload["recommended_path"] = skill_recommendation(skill, row.get("agentifiability", ""))
                refs["skill_feasibility"] = add(EvidenceRecord(
                    field_name="skill_feasibility",
                    value=skill_payload,
                    evidence_class=EvidenceClass.CONSENSUS if is_provisional else EvidenceClass.SUPPLEMENTAL,
                    status=EvidenceStatus.UNVERIFIED if is_provisional else EvidenceStatus.ACTIVE,
                    confidence="LOW" if is_provisional else "HIGH",
                    conflict_note="基于业务方口头确认，待补01层书面材料" if is_provisional else "",
                    source=SourceRef(
                        source_system="OB知识库",
                        source_object="L4流程_Skill封装可行性评估_确认最终版_v2.xlsx",
                        source_key=f"{source_code}@row:{source_row}",
                        source_field="L4明细_最终确认版:A:N",
                        source_version="v2",
                    ),
                ))
            category = self.l3_position_category.get(l3_code)
            position_family = None
            if category is not None:
                position_family = {
                    "family_code": category["family_code"],
                    "family_name": category["family_name"],
                    "category_name": category["category_name"],
                    "category_type": category["category_type"],
                }
                refs["position_family"] = add(EvidenceRecord(
                    field_name="position_family",
                    value=position_family,
                    evidence_class=EvidenceClass.SUPPLEMENTAL,
                    status=EvidenceStatus.ACTIVE,
                    source=SourceRef(
                        source_system="EA项目_权威数据",
                        source_object="2026-07-20_68L3岗位族归属设计_v6.1_SUBMITTED.md",
                        source_key=l3_code,
                        source_field="各族详细映射",
                        source_version="v6.1",
                    ),
                ))
            business_evidence = [
                {
                    "schema": item["schema"],
                    "table": item["table"],
                    "row_count": self.business_table_counts.get(f"{item['schema']}.{item['table']}"),
                    "matched_columns": item["matched_columns"],
                    "rationale": item["rationale"],
                    "confidence": item["confidence"],
                    "evidence_type": item["evidence_type"],
                }
                for item in self.business_table_map.get(code, [])
            ]
            if business_evidence:
                refs["business_evidence"] = add(EvidenceRecord(
                    field_name="business_evidence",
                    value=business_evidence,
                    evidence_class=EvidenceClass.CONSENSUS,
                    status=EvidenceStatus.UNVERIFIED,
                    confidence="LOW",
                    conflict_note="人工按业务含义匹配到业务数据仓库表，非外键关联，未经业务方确认",
                    source=SourceRef(
                        source_system="业务数据仓库",
                        source_object=", ".join(f"{i['schema']}.{i['table']}" for i in business_evidence),
                        source_key=code,
                        source_field="人工匹配",
                        source_version="2026-08-04试点",
                    ),
                ))
            l4s.append({
                "l4_code": code,
                "l4_name": row.get("l4_name", ""),
                "deliverable": row.get("l4_deliverable", ""),
                "deliverable_type": row.get("l4_deliverable_type", ""),
                "tier": row.get("agentifiability", ""),
                "human_touchpoint": row.get("agent_human_touchpoint", ""),
                "d1_d6": d1_d6,
                "skill_feasibility": skill_payload,
                "position_family": position_family,
                "business_evidence": business_evidence,
                "evidence_refs": refs,
            })

        vn_items = []
        for row in nodes:
            key = row.get("vn_id", "")
            refs = {
                field: add(authoritative(field, row.get(field), "dim_vn", key, field))
                for field in (
                    "vn_name", "overall_judgment", "is_fused", "priority",
                    "gate1_data_linked", "gate2_grounded", "gate3_traceable",
                )
            }
            vn_items.append({
                "vn_id": key,
                "vn_name": row.get("vn_name", ""),
                "is_fused": row.get("is_fused"),
                "priority": row.get("priority", ""),
                "overall_judgment": row.get("overall_judgment", ""),
                "evidence_refs": refs,
            })

        if supplemental:
            for record in supplemental:
                if record.evidence_class is not EvidenceClass.SUPPLEMENTAL:
                    raise ValueError("supplemental参数只接受SUPPLEMENTAL证据")
                add(record)

        l4_codes = {item["l4_code"] for item in l4s}
        vn_ids = {item["vn_id"] for item in vn_items}
        linked_sops = [
            row for row in self.sop_records
            if l3_code in " ".join(str(value) for value in row.values())
            or row.get("sop_ref", "") in vn_ids
        ]
        linked_rules = [row for row in self.rule_records if row.get("node_id", "") in vn_ids]
        for row in linked_sops:
            row["evidence_ref"] = add(EvidenceRecord(
                field_name=f"readiness.sop.{row.get('sop_ref', '')}",
                value={"title": row.get("sop_title", ""), "file": row.get("sop_file", "")},
                evidence_class=EvidenceClass.SUPPLEMENTAL,
                status=EvidenceStatus.ACTIVE,
                source=SourceRef(
                    source_system="VNW数据基础",
                    source_object="T19_SOP生产进度_全域_v2.0.csv",
                    source_key=row.get("sop_ref", ""),
                    source_field="sop_file",
                    source_version=row.get("sop_version", ""),
                ),
            ))
        for row in linked_rules:
            row["evidence_ref"] = add(EvidenceRecord(
                field_name=f"readiness.rule.{row.get('rule_id', '')}",
                value={
                    "name": row.get("rule_name", ""),
                    "action": row.get("rule_action", ""),
                    "standard": row.get("rule_standard", ""),
                },
                evidence_class=EvidenceClass.SUPPLEMENTAL,
                status=EvidenceStatus.ACTIVE,
                source=SourceRef(
                    source_system="VNW数据基础",
                    source_object="T5_规则清单_全域_v3.0.csv",
                    source_key=row.get("rule_id", ""),
                    source_field="rule_action",
                    source_version=row.get("version", ""),
                ),
            ))

        has_l4 = bool(l4s)
        names_complete = has_l4 and all(item["l4_name"] for item in l4s)
        blueprint_parsed = bool(parsed_blueprint and parsed_blueprint["structure_status"] == "PARSED" and parsed_blueprint["steps"])
        has_sop = bool(linked_sops)
        has_rules = bool(linked_rules)
        deliverable_count = sum(bool(item["deliverable"]) for item in l4s)
        deliverable_ratio = deliverable_count / len(l4s) if l4s else 0
        d_complete_count = sum(all(item["d1_d6"][field] is not None for field in D_FIELDS) for item in l4s)
        d_complete = has_l4 and d_complete_count == len(l4s)
        skill_verified_count = sum(
            bool(item.get("skill_feasibility"))
            and item["skill_feasibility"].get("verification_status") == "VERIFIED"
            for item in l4s
        )
        skill_provisional_count = sum(
            bool(item.get("skill_feasibility"))
            and item["skill_feasibility"].get("verification_status") == "PROVISIONAL"
            for item in l4s
        )
        gate_m_checks = [
            {"rule_id": "M-001", "passed": has_l4, "detail": f"L4数量={len(l4s)}"},
            {"rule_id": "M-002", "passed": blueprint_parsed, "detail": blueprint.filename if blueprint_parsed else "缺少可解析流程蓝图"},
            {"rule_id": "M-003", "passed": names_complete, "detail": "L4名称完整" if names_complete else "L4名称缺失"},
            {"rule_id": "M-004", "passed": deliverable_ratio >= 0.7, "detail": f"交付物完整={deliverable_count}/{len(l4s)}（最低门槛70%）"},
            {"rule_id": "M-005", "passed": d_complete, "detail": f"D1-D6完整={d_complete_count}/{len(l4s)}"},
        ]
        gate_m = "PASS" if all(c["passed"] for c in gate_m_checks) else "BLOCKED"

        tiers_valid = has_l4 and all(item["tier"] in VALID_TIERS for item in l4s)
        blueprint_task_l4s = {
            code for step in (parsed_blueprint or {}).get("steps", [])
            for code in step.get("l4_codes", []) if code in l4_codes
        }
        sop_task_l4s = {
            code for row in linked_sops for code in row.get("l4_codes", [])
            if code in l4_codes
        }
        analysis_candidate = self.analysis_packages.get(l3_code)
        candidate_l4_codes = {
            item.get("l4_code", "")
            for item in (analysis_candidate or {}).get("l4_analysis", [])
            if item.get("l4_code")
        }
        analysis_candidate_current = bool(analysis_candidate) and candidate_l4_codes == l4_codes
        analyzed_task_l4s = {
            task.get("l4_code", "") for task in (analysis_candidate or {}).get("tasks", [])
            if analysis_candidate_current
            and task.get("evidence_refs")
            and task.get("l4_code") in l4_codes
        }
        task_l4s = blueprint_task_l4s | sop_task_l4s | analyzed_task_l4s
        task_ratio = len(task_l4s) / len(l4s) if l4s else 0
        mapped_by_vn: dict[str, set[str]] = {}
        for mapping in mappings:
            mapped_by_vn.setdefault(mapping.get("vn_id", ""), set()).update(
                expand_l4_mapping(mapping.get("l4_code", ""), l4_codes)
            )
        rule_l4s = {
            code for row in linked_rules for code in mapped_by_vn.get(row.get("node_id", ""), set())
        }
        critical_vns = {
            item["vn_id"] for item in vn_items
            if item.get("is_fused") is True or str(item.get("priority", "")).upper() in {"P0", "HIGH"}
        }
        critical_l4s = {code for vn_id in critical_vns for code in mapped_by_vn.get(vn_id, set())}
        critical_covered = not critical_l4s or critical_l4s <= task_l4s
        gate_e_checks = [
            {"rule_id": "E-001", "passed": gate_m == "PASS", "detail": f"Gate M={gate_m}"},
            {"rule_id": "E-002", "passed": deliverable_ratio >= 0.9, "detail": f"交付物完整={deliverable_count}/{len(l4s)}（完整门槛90%）"},
            {"rule_id": "E-003", "passed": tiers_valid, "detail": "Tier合法" if tiers_valid else "Tier缺失或非法"},
            {"rule_id": "E-004", "passed": task_ratio >= 0.8, "detail": f"可追溯任务覆盖={len(task_l4s)}/{len(l4s)}（完整门槛80%）"},
            {"rule_id": "E-005", "passed": critical_covered, "detail": f"关键L4任务覆盖={len(critical_l4s & task_l4s)}/{len(critical_l4s)}"},
            {"rule_id": "E-006", "passed": bool(mappings), "detail": f"价值节点-L4映射={len(mappings)}"},
            {"rule_id": "E-007", "passed": bool(rule_l4s), "detail": f"规则可定位L4={len(rule_l4s)}/{len(l4s)}"},
            {"rule_id": "E-008", "passed": has_sop, "detail": f"SOP/任务执行材料={len(linked_sops)}份" if has_sop else "缺少SOP草稿或同等任务执行材料（不阻断建模）"},
            {"rule_id": "E-009", "passed": has_rules, "detail": f"有效规则={len(linked_rules)}条" if has_rules else "缺少可关联的规则记录（不阻断建模）"},
        ]
        minimum_evidence = (
            gate_m == "PASS"
        )
        gate_e = (
            "PASS" if all(c["passed"] for c in gate_e_checks)
            else "CONDITIONAL" if minimum_evidence
            else "BLOCKED"
        )

        mapping_complete = bool(mappings)
        not_fused = bool(nodes) and all(row.get("is_fused") is False for row in nodes)
        gate_a_checks = [
            {"rule_id": "A-001", "passed": gate_e in {"PASS", "CONDITIONAL"}, "detail": f"Gate E={gate_e}"},
            {"rule_id": "A-002", "passed": d_complete, "detail": f"D1-D6完整={d_complete_count}/{len(l4s)}"},
            {"rule_id": "A-003", "passed": mapping_complete, "detail": f"价值节点-L4映射={len(mappings)}"},
            {"rule_id": "A-004", "passed": not_fused, "detail": "无熔断节点" if not_fused else "存在熔断或无价值节点"},
        ]
        gate_a = "PASS" if all(c["passed"] for c in gate_a_checks) else "BLOCKED"

        l3_name = processes[0].get("l3_name", "") if processes else (blueprint.l3_name if blueprint else "")
        d1d6_rows = []
        skill_rows = []
        for l4 in l4s:
            supplement = self.d1d6_supplement.get(l4["l4_code"])
            if supplement is None:
                supplement = self.d1d6_supplement.get(
                    d1d6_name_key(l3_code, l4.get("l4_name", ""))
                )
            if supplement and supplement.get("_source_row"):
                d1d6_rows.append(supplement["_source_row"])
            skill = self.skill_feasibility.get(l4["l4_code"])
            if skill is None:
                skill = self.skill_feasibility.get(
                    d1d6_name_key(l3_code, l4.get("l4_name", ""))
                )
            if skill and skill.get("_source_row"):
                skill_rows.append(skill["_source_row"])
        blueprint_lines = []
        if parsed_blueprint:
            for collection in ("steps", "decisions", "blueprint_value_nodes", "raci"):
                blueprint_lines.extend(
                    item.get("source_line")
                    for item in parsed_blueprint.get(collection, [])
                    if item.get("source_line")
                )
        source_locations = {
            "process_analytics.dim_process": {
                "record_keys": sorted({item["l4_code"] for item in l4s}),
            },
            "process_analytics.dim_vn": {
                "record_keys": sorted({item["vn_id"] for item in vn_items}),
            },
            "L4两阶段复核_全量368条_合并版_v1.0.csv": {
                "rows": sorted(set(d1d6_rows)),
            },
            "L4流程_Skill封装可行性评估_确认最终版_v2.xlsx": {
                "rows": sorted(set(skill_rows)),
            },
            "T5_规则清单_全域_v3.0.csv": {
                "rows": sorted({
                    row.get("_source_row") for row in linked_rules
                    if row.get("_source_row")
                }),
            },
            "T19_SOP生产进度_全域_v2.0.csv": {
                "rows": sorted({
                    row.get("_source_row") for row in linked_sops
                    if row.get("_source_row")
                }),
            },
            "2026-07-20_68L3岗位族归属设计_v6.1_SUBMITTED.md": {
                "record_keys": [l3_code] if l3_code in self.l3_position_category else [],
            },
        }
        if blueprint:
            source_locations[blueprint.filename] = {
                "lines": sorted(set(blueprint_lines)),
            }
        model = {
            "schema_version": self.schema_version,
            "l3_code": l3_code,
            "l3_name": l3_name,
            "has_demo": l3_code in self.demo_registry,
            "demo_file": self.demo_registry.get(l3_code, ""),
            "source_policy": {
                "database_authority": "process_analytics",
                "supplemental_requires_active": True,
                "consensus_separate": True,
                "reverse_writeback": False,
            },
            "source_locations": source_locations,
            "blueprint": {
                "coverage": "INDEXED" if blueprint else "MISSING",
                "version": blueprint.version if blueprint else "",
                "filename": blueprint.filename if blueprint else "",
                "structure_status": parsed_blueprint["structure_status"] if parsed_blueprint else ("INDEX_ONLY" if blueprint else "UNAVAILABLE"),
                "steps": parsed_blueprint["steps"] if parsed_blueprint else [],
                "decisions": parsed_blueprint["decisions"] if parsed_blueprint else [],
                "edges": parsed_blueprint["edges"] if parsed_blueprint else [],
                "blueprint_value_nodes": parsed_blueprint["blueprint_value_nodes"] if parsed_blueprint else [],
                "raci": parsed_blueprint["raci"] if parsed_blueprint else [],
                "source_path": parsed_blueprint["source_path"] if parsed_blueprint else "",
                "source_hash": parsed_blueprint["source_hash"] if parsed_blueprint else "",
                "diagnostics": parsed_blueprint["diagnostics"] if parsed_blueprint else {},
                "note": (
                    "已从蓝图正文解析；步骤顺序边由显式步骤编号派生，判断分支保留原文行号"
                    if parsed_blueprint and parsed_blueprint["structure_status"] == "PARSED"
                    else "正文缺少可确认结构或与当前数据库冲突，禁止生成箭头、决策点和回退关系"
                ),
            },
            "l2_capabilities": l2s,
            "l4s": l4s,
            "value_nodes": vn_items,
            "vn_l4_mappings": mappings,
            "kpi_mappings": kpis,
            "value_stream_mappings": value_stages,
            "gates": {
                "M": gate_result(gate_m, gate_m_checks),
                "E": gate_result(gate_e, gate_e_checks),
                "A": gate_result(gate_a, gate_a_checks),
            },
            "model_readiness": {
                "status": (
                    "FULL_MODEL" if gate_m == "PASS" and gate_e == "PASS"
                    else "LIMITED_MODEL" if gate_m == "PASS" and gate_e == "CONDITIONAL"
                    else "WAITING_INPUT"
                ),
                "model_generation_allowed": gate_m == "PASS" and gate_e in {"PASS", "CONDITIONAL"},
                "thresholds": {
                    "deliverable_full": 0.9,
                    "deliverable_minimum": 0.7,
                    "task_full": 0.8,
                    "task_minimum": 0.6,
                    "critical_l4_required": 1.0,
                },
                "coverage": {
                    "sop_count": len(linked_sops),
                    "rule_count": len(linked_rules),
                    "deliverable": {"covered": deliverable_count, "total": len(l4s)},
                    "task": {"covered": len(task_l4s), "total": len(l4s)},
                    "rule_l4": {"covered": len(rule_l4s), "total": len(l4s)},
                    "critical_task": {"covered": len(critical_l4s & task_l4s), "total": len(critical_l4s)},
                    "skill_feasibility": {
                        "verified": skill_verified_count,
                        "provisional": skill_provisional_count,
                        "total": len(l4s),
                    },
                },
                "linked_sources": {
                    "sops": [
                        {"ref": row.get("sop_ref", ""), "file": row.get("sop_file", ""), "evidence_ref": row.get("evidence_ref", "")}
                        for row in linked_sops
                    ],
                    "rules": [
                        {"rule_id": row.get("rule_id", ""), "node_id": row.get("node_id", ""), "name": row.get("rule_name", ""), "evidence_ref": row.get("evidence_ref", "")}
                        for row in linked_rules
                    ],
                },
            },
            "evidence_registry": sorted(evidence.values(), key=lambda item: item["evidence_id"]),
            "supplemental_evidence_refs": [record.evidence_id for record in (supplemental or [])],
        }
        analysis_evidence_ids = eligible_analysis_evidence_ids(evidence)
        fallback_analysis = build_analysis_envelope(
            l3_code=l3_code, l4s=l4s, blueprint=model["blueprint"], evidence_ids=analysis_evidence_ids,
        )
        model["analysis"] = analysis_candidate or fallback_analysis
        try:
            validate_analysis_package(
                model["analysis"],
                evidence_ids=analysis_evidence_ids,
                l4_codes={item["l4_code"] for item in l4s},
            )
        except ValueError as exc:
            if not analysis_candidate:
                raise
            model["stale_analysis"] = {
                "status": "ANALYSIS_INPUT_CHANGED",
                "reason": str(exc),
                "previous_analysis_status": analysis_candidate.get("analysis_status", ""),
                "previous_input_snapshot_hash": (analysis_candidate.get("model_run") or {}).get("input_snapshot_hash", ""),
            }
            model["analysis"] = fallback_analysis
            validate_analysis_package(
                model["analysis"],
                evidence_ids=analysis_evidence_ids,
                l4_codes={item["l4_code"] for item in l4s},
            )
        model["analysis_input_hash"] = analysis_input_hash(model)
        run_hash = (model["analysis"].get("model_run") or {}).get("input_snapshot_hash", "")
        if model.get("stale_analysis") or (run_hash and run_hash != model["analysis_input_hash"]):
            model["analysis_freshness"] = "INPUT_CHANGED"
        elif model["analysis"].get("analysis_status") == "PENDING_MODEL":
            model["analysis_freshness"] = "PENDING_MODEL"
        elif run_hash == model["analysis_input_hash"]:
            model["analysis_freshness"] = "CURRENT"
        else:
            model["analysis_freshness"] = "UNVERSIONED_REVIEWED_BASELINE"
        model["unified_analysis"] = l3_synthesis(model, self.analysis_confirmations_dir)
        return model

    def build_and_write(
        self,
        l3_codes: list[str],
        output_dir: Path,
        supplemental_by_l3: dict[str, list[EvidenceRecord]] | None = None,
    ) -> list[dict]:
        results = []
        models = []
        for code in l3_codes:
            try:
                model = self.build(code, (supplemental_by_l3 or {}).get(code, []))
            except (KeyError, TypeError, ValueError) as exc:
                raise type(exc)(f"{code}快照构建失败：{exc}") from exc
            models.append(model)
            results.append(write_snapshot(model, output_dir))
        indexed_models = {}
        for snapshot_path in output_dir.glob("L3-*.json"):
            if snapshot_path.name.endswith(".manifest.json"):
                continue
            stored_model = json.loads(snapshot_path.read_text(encoding="utf-8"))
            if stored_model.get("schema_version") == self.schema_version:
                indexed_models[stored_model["l3_code"]] = stored_model
        for model in models:
            indexed_models[model["l3_code"]] = model
        index = {
            "schema_version": "vnw.l3-model-index.v1",
            "source_policy": "PostgreSQL权威；仅生效OB材料可补充；工作坊判断仅存浏览器本地",
            "models": [
                {
                    "l3_code": model["l3_code"],
                    "l3_name": model["l3_name"],
                    "l2_capabilities": [row["l2_name"] for row in model["l2_capabilities"]],
                    "value_streams": [
                        {
                            "vs_code": row.get("vs_code", ""),
                            "vs_name": row.get("vs_name", ""),
                            "stage_code": row.get("stage_code", ""),
                            "stage_name": row.get("stage_name", ""),
                            "stage_sequence": row.get("stage_sequence"),
                        }
                        for row in model["value_stream_mappings"]
                    ],
                    "kpis": [
                        {"kpi_name": row["kpi_name"], "source_type": row["source_type"]}
                        for row in model["kpi_mappings"]
                    ],
                    "l4_count": len(model["l4s"]),
                    "value_node_count": len(model["value_nodes"]),
                    "blueprint_coverage": model["blueprint"]["coverage"],
                    "blueprint_version": model["blueprint"]["version"],
                    "gates": {key: value["status"] for key, value in model["gates"].items()},
                    "classification": model["model_readiness"]["status"],
                    "model_generation_allowed": model["model_readiness"]["model_generation_allowed"],
                    "readiness_coverage": model["model_readiness"]["coverage"],
                    "analysis_status": model["analysis"]["analysis_status"],
                    "analysis_input_hash": model.get("analysis_input_hash", ""),
                    "analysis_run_input_hash": (model["analysis"].get("model_run") or {}).get("input_snapshot_hash", ""),
                    "analysis_run_at": (model["analysis"].get("model_run") or {}).get("generated_at"),
                    "production_status": (
                        "BLOCKED_INPUT" if not model["model_readiness"]["model_generation_allowed"]
                        else "ANALYSIS_INPUT_CHANGED" if model.get("stale_analysis")
                        else "RUN_PREPARED" if model["l3_code"] in self.prepared_analysis_codes and model["analysis"]["analysis_status"] == "PENDING_MODEL"
                        else "READY_TO_RUN" if model["analysis"]["analysis_status"] == "PENDING_MODEL"
                        else "REVIEWED_BASELINE" if model["analysis"]["analysis_status"] == "REVIEWED" and not (model["analysis"].get("model_run") or {}).get("input_snapshot_hash")
                        else "ANALYSIS_CURRENT" if (model["analysis"].get("model_run") or {}).get("input_snapshot_hash") == model.get("analysis_input_hash", "")
                        else "ANALYSIS_INPUT_CHANGED"
                    ),
                    "highest_gate": (
                        "A" if model["gates"]["A"]["status"] == "PASS"
                        else "E" if model["gates"]["E"]["status"] in {"PASS", "CONDITIONAL"}
                        else "M" if model["gates"]["M"]["status"] == "PASS"
                        else "NONE"
                    ),
                    "gap_reasons": [
                        check["detail"]
                        for gate in ("M", "E", "A")
                        for check in model["gates"][gate]["checks"]
                        if not check["passed"]
                    ],
                    "blueprint_structure_status": model["blueprint"]["structure_status"],
                    "snapshot_file": f"{model['l3_code']}.json",
                    "has_demo": model.get("has_demo", False),
                    "demo_file": model.get("demo_file", ""),
                }
                for model in sorted(indexed_models.values(), key=lambda item: item["l3_code"])
            ],
        }
        production_order = sorted(
            (
                item for item in index["models"]
                if item["production_status"] == "READY_TO_RUN"
            ),
            key=lambda item: (
                0 if item["classification"] == "FULL_MODEL" else 1,
                len(item["gap_reasons"]),
                item["l4_count"],
                item["l3_code"],
            ),
        )
        index["production_summary"] = {
            status: sum(item["production_status"] == status for item in index["models"])
            for status in ("READY_TO_RUN", "RUN_PREPARED", "ANALYSIS_CURRENT", "ANALYSIS_INPUT_CHANGED", "REVIEWED_BASELINE", "BLOCKED_INPUT")
        }
        index["recommended_batch"] = [
            {"l3_code": item["l3_code"], "l3_name": item["l3_name"], "classification": item["classification"], "l4_count": item["l4_count"]}
            for item in production_order[:3]
        ]
        index["prepared_batch"] = [
            {"l3_code": item["l3_code"], "l3_name": item["l3_name"], "classification": item["classification"], "l4_count": item["l4_count"]}
            for item in index["models"] if item["production_status"] == "RUN_PREPARED"
        ]
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "index.json").write_text(
            json.dumps(index, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        return results
